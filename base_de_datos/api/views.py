from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.pagination import PageNumberPagination
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.db import DatabaseError
from django.db.models import Q
from datetime import datetime
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from ..models import RegistroVehiculo
from clientes.models import Cliente
from .permissions import RolePermission, ModulePermission


def serialize_registro(registro):
    """Convierte un objeto RegistroVehiculo a diccionario"""
    return {
        'id': registro.id,
        'usuario': {
            'id': registro.usuario.id,
            'name': f"{registro.usuario.first_name} {registro.usuario.last_name}".strip(),
        } if registro.usuario else None,
        'cliente': {
            'id': registro.cliente.id,
            'nombre': registro.cliente.nombre,
        } if registro.cliente else None,
        # Trámite
        'tipo_tramite': registro.tipo_tramite,
        'tipo_vehiculo': registro.tipo_vehiculo,
        # Titular
        'es_propietario': registro.es_propietario,
        'tipo_documento': registro.tipo_documento,
        'numero_documento': registro.numero_documento,
        'nombre_completo': registro.nombre_completo,
        'telefono': registro.telefono,
        # Vehículo
        'placa': registro.placa,
        'clase': registro.clase,
        'clasificacion': registro.clasificacion,
        'tipo_servicio': registro.tipo_servicio,
        'tipo_carroceria': registro.tipo_carroceria,
        'vin': registro.vin,
        'num_motor': registro.num_motor,
        'num_chasis': registro.num_chasis,
        'marca': registro.marca,
        'linea': registro.linea,
        'modelo': registro.modelo,
        'cilindraje': registro.cilindraje,
        'color': registro.color,
        'organismo_transito': registro.organismo_transito,
        # Cotización
        'precio_lay': str(registro.precio_lay) if registro.precio_lay is not None else None,
        'comision': str(registro.comision) if registro.comision is not None else None,
        # Timestamps
        'created_at': registro.created_at,
        'updated_at': registro.updated_at,
        'deleted_at': registro.deleted_at,
    }


@api_view(['POST'])
@permission_classes([IsAuthenticated, RolePermission(['admin', 'SuperAdmin', 'vendedor']), ModulePermission('base_de_datos', 'create')])
def create_registro(request):
    """Crear un nuevo registro de vehículo"""
    try:
        required_fields = ['cliente', 'numero_documento', 'nombre_completo']
        for field in required_fields:
            if not request.data.get(field):
                return Response(
                    {"error": f"El campo {field} es requerido."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Validar que el cliente exista
        cliente_id = request.data.get('cliente')
        try:
            cliente = Cliente.objects.get(pk=cliente_id)
            if cliente.deleted_at is not None:
                return Response(
                    {"error": "El cliente especificado está eliminado."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Cliente.DoesNotExist:
            return Response(
                {"error": "El cliente especificado no existe."},
                status=status.HTTP_404_NOT_FOUND
            )

        registro = RegistroVehiculo.objects.create(
            usuario=request.user,
            cliente=cliente,
            # Trámite
            tipo_tramite=request.data.get('tipo_tramite', 'SOAT'),
            tipo_vehiculo=request.data.get('tipo_vehiculo', 'USADO'),
            # Titular
            es_propietario=request.data.get('es_propietario', True),
            tipo_documento=request.data.get('tipo_documento', 'C'),
            numero_documento=request.data.get('numero_documento'),
            nombre_completo=request.data.get('nombre_completo'),
            telefono=request.data.get('telefono', ''),
            # Vehículo
            placa=request.data.get('placa', ''),
            clase=request.data.get('clase', ''),
            clasificacion=request.data.get('clasificacion', ''),
            tipo_servicio=request.data.get('tipo_servicio', ''),
            tipo_carroceria=request.data.get('tipo_carroceria', ''),
            vin=request.data.get('vin', ''),
            num_motor=request.data.get('num_motor', ''),
            num_chasis=request.data.get('num_chasis', ''),
            marca=request.data.get('marca', ''),
            linea=request.data.get('linea', ''),
            modelo=request.data.get('modelo', ''),
            cilindraje=request.data.get('cilindraje', ''),
            color=request.data.get('color', ''),
            organismo_transito=request.data.get('organismo_transito', ''),
            precio_lay=request.data.get('precio_lay') or None,
            comision=request.data.get('comision') or None,
        )

        return Response(serialize_registro(registro), status=status.HTTP_201_CREATED)

    except DatabaseError as e:
        return Response(
            {"error": f"Error de base de datos: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    except Exception as e:
        return Response(
            {"error": f"Error inesperado: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated, ModulePermission('base_de_datos', 'view')])
def list_registros(request):
    """Listar registros de vehículos con filtros y paginación"""
    try:
        registros = RegistroVehiculo.objects.select_related(
            'usuario', 'cliente'
        ).all()

        # Filtro de búsqueda general
        search_query = request.query_params.get('search', None)
        if search_query:
            registros = registros.filter(
                Q(placa__icontains=search_query) |
                Q(nombre_completo__icontains=search_query) |
                Q(numero_documento__icontains=search_query) |
                Q(marca__icontains=search_query) |
                Q(linea__icontains=search_query) |
                Q(vin__icontains=search_query) |
                Q(num_chasis__icontains=search_query)
            )

        # Filtro por cliente
        cliente_id = request.query_params.get('cliente', None)
        if cliente_id:
            registros = registros.filter(cliente_id=cliente_id)

        # Filtro por tipo de trámite
        tipo_tramite = request.query_params.get('tipo_tramite', None)
        if tipo_tramite:
            registros = registros.filter(tipo_tramite=tipo_tramite)

        # Filtro por tipo de vehículo
        tipo_vehiculo = request.query_params.get('tipo_vehiculo', None)
        if tipo_vehiculo:
            registros = registros.filter(tipo_vehiculo=tipo_vehiculo)

        # Filtro por tipo de documento
        tipo_documento = request.query_params.get('tipo_documento', None)
        if tipo_documento:
            registros = registros.filter(tipo_documento=tipo_documento)

        # Filtro por es_propietario
        es_propietario = request.query_params.get('es_propietario', None)
        if es_propietario is not None:
            registros = registros.filter(es_propietario=es_propietario == '1')

        # Filtro por usuario
        usuario_id = request.query_params.get('usuario', None)
        if usuario_id:
            registros = registros.filter(usuario_id=usuario_id)

        # Filtro por fecha de creación
        start_date_str = request.query_params.get('start_date', None)
        end_date_str = request.query_params.get('end_date', None)

        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                registros = registros.filter(created_at__gte=start_date)
            except ValueError:
                return Response(
                    {"error": "El formato de start_date debe ser YYYY-MM-DD."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                end_date_inclusive = datetime.combine(end_date, datetime.max.time())
                registros = registros.filter(created_at__lte=end_date_inclusive)
            except ValueError:
                return Response(
                    {"error": "El formato de end_date debe ser YYYY-MM-DD."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        # Filtro para incluir eliminados
        include_deleted = request.query_params.get('include_deleted', None)
        if include_deleted != '1':
            registros = registros.filter(deleted_at__isnull=True)

        # Ordenar
        registros = registros.order_by('-created_at')

        # Paginación
        page_size_param = request.query_params.get('page_size', 10)
        try:
            page_size_int = int(page_size_param)
        except (ValueError, TypeError):
            page_size_int = 10

        paginator = PageNumberPagination()
        paginator.page_size = page_size_int
        paginated_registros = paginator.paginate_queryset(registros, request)

        data = [serialize_registro(r) for r in paginated_registros]
        return paginator.get_paginated_response(data)

    except Exception as e:
        return Response(
            {"error": f"Error al obtener registros: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated, ModulePermission('base_de_datos', 'view')])
def get_registro(request, pk):
    """Obtener un registro de vehículo por ID"""
    try:
        registro = get_object_or_404(
            RegistroVehiculo.objects.select_related('usuario', 'cliente'),
            pk=pk
        )
        return Response(serialize_registro(registro), status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            {"error": f"Error al obtener registro: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['PUT'])
@permission_classes([IsAuthenticated, RolePermission(['admin', 'SuperAdmin', 'vendedor']), ModulePermission('base_de_datos', 'edit')])
def update_registro(request, pk):
    """Actualizar un registro de vehículo"""
    try:
        registro = get_object_or_404(RegistroVehiculo, pk=pk)

        # Validar cliente si se proporciona
        if 'cliente' in request.data:
            cliente_id = request.data.get('cliente')
            try:
                cliente = Cliente.objects.get(pk=cliente_id)
                if cliente.deleted_at is not None:
                    return Response(
                        {"error": "El cliente especificado está eliminado."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                registro.cliente = cliente
            except Cliente.DoesNotExist:
                return Response(
                    {"error": "El cliente especificado no existe."},
                    status=status.HTTP_404_NOT_FOUND
                )

        # Campos actualizables
        campos = [
            'tipo_tramite', 'tipo_vehiculo', 'es_propietario',
            'tipo_documento', 'numero_documento', 'nombre_completo', 'telefono',
            'placa', 'clase', 'clasificacion', 'tipo_servicio', 'tipo_carroceria',
            'vin', 'num_motor', 'num_chasis', 'marca', 'linea', 'modelo',
            'cilindraje', 'color', 'organismo_transito',
            'precio_lay', 'comision',
        ]

        for campo in campos:
            if campo in request.data:
                setattr(registro, campo, request.data.get(campo))

        registro.save()

        return Response(serialize_registro(registro), status=status.HTTP_200_OK)

    except DatabaseError as e:
        return Response(
            {"error": f"Error de base de datos: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
    except Exception as e:
        return Response(
            {"error": f"Error inesperado: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['DELETE'])
@permission_classes([IsAuthenticated, RolePermission(['admin', 'SuperAdmin']), ModulePermission('base_de_datos', 'delete')])
def delete_registro(request, pk):
    """Eliminar un registro de vehículo (soft delete)"""
    try:
        registro = get_object_or_404(RegistroVehiculo, pk=pk)
        registro.soft_delete()
        return Response(
            {"message": "Registro eliminado correctamente"},
            status=status.HTTP_200_OK
        )
    except Exception as e:
        return Response(
            {"error": f"Error al eliminar registro: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([IsAuthenticated, RolePermission(['admin', 'SuperAdmin']), ModulePermission('base_de_datos', 'delete')])
def restore_registro(request, pk):
    """Restaurar un registro de vehículo eliminado"""
    try:
        registro = get_object_or_404(RegistroVehiculo, pk=pk)
        if not registro.is_deleted:
            return Response(
                {"error": "El registro no está eliminado"},
                status=status.HTTP_400_BAD_REQUEST
            )
        registro.restore()
        return Response(serialize_registro(registro), status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            {"error": f"Error al restaurar registro: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['DELETE'])
@permission_classes([IsAuthenticated, RolePermission(['admin', 'SuperAdmin']), ModulePermission('base_de_datos', 'delete')])
def hard_delete_registro(request, pk):
    """Eliminar permanentemente un registro de vehículo"""
    try:
        registro = get_object_or_404(RegistroVehiculo, pk=pk)
        registro.delete()
        return Response(
            {"message": "Registro eliminado permanentemente"},
            status=status.HTTP_204_NO_CONTENT
        )
    except Exception as e:
        return Response(
            {"error": f"Error al eliminar registro: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated, ModulePermission('base_de_datos', 'view')])
def export_registros_excel(request):
    """Exportar registros de vehículos a Excel con los mismos filtros del listado"""
    try:
        registros = RegistroVehiculo.objects.select_related(
            'usuario', 'cliente'
        ).all()

        # ── Aplicar los mismos filtros que list_registros ──
        search_query = request.query_params.get('search', None)
        if search_query:
            registros = registros.filter(
                Q(placa__icontains=search_query) |
                Q(nombre_completo__icontains=search_query) |
                Q(numero_documento__icontains=search_query) |
                Q(marca__icontains=search_query) |
                Q(linea__icontains=search_query) |
                Q(vin__icontains=search_query) |
                Q(num_chasis__icontains=search_query)
            )

        cliente_id = request.query_params.get('cliente', None)
        if cliente_id:
            registros = registros.filter(cliente_id=cliente_id)

        tipo_tramite = request.query_params.get('tipo_tramite', None)
        if tipo_tramite:
            registros = registros.filter(tipo_tramite=tipo_tramite)

        tipo_vehiculo = request.query_params.get('tipo_vehiculo', None)
        if tipo_vehiculo:
            registros = registros.filter(tipo_vehiculo=tipo_vehiculo)

        tipo_documento = request.query_params.get('tipo_documento', None)
        if tipo_documento:
            registros = registros.filter(tipo_documento=tipo_documento)

        es_propietario = request.query_params.get('es_propietario', None)
        if es_propietario is not None:
            registros = registros.filter(es_propietario=es_propietario == '1')

        usuario_id = request.query_params.get('usuario', None)
        if usuario_id:
            registros = registros.filter(usuario_id=usuario_id)

        start_date_str = request.query_params.get('start_date', None)
        end_date_str = request.query_params.get('end_date', None)

        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                registros = registros.filter(created_at__gte=start_date)
            except ValueError:
                return Response(
                    {"error": "El formato de start_date debe ser YYYY-MM-DD."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                end_date_inclusive = datetime.combine(end_date, datetime.max.time())
                registros = registros.filter(created_at__lte=end_date_inclusive)
            except ValueError:
                return Response(
                    {"error": "El formato de end_date debe ser YYYY-MM-DD."},
                    status=status.HTTP_400_BAD_REQUEST
                )

        include_deleted = request.query_params.get('include_deleted', None)
        if include_deleted != '1':
            registros = registros.filter(deleted_at__isnull=True)

        registros = registros.order_by('-created_at')

        # ── Generar el archivo Excel ──
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Registros de Vehículos'

        # Estilos
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='1976D2', end_color='1976D2', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0'),
        )
        cell_alignment = Alignment(vertical='center', wrap_text=False)
        date_format = 'DD/MM/YYYY HH:MM'

        # Título del reporte
        ws.merge_cells('A1:T1')
        title_cell = ws['A1']
        title_cell.value = 'Reporte de Base de Datos - Registros de Vehículos'
        title_cell.font = Font(name='Calibri', bold=True, size=14, color='1976D2')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        # Subtítulo con filtros aplicados y fecha de generación
        filtros_texto = []
        if search_query:
            filtros_texto.append(f'Búsqueda: "{search_query}"')
        if tipo_tramite:
            filtros_texto.append(f'Trámite: {tipo_tramite}')
        if tipo_vehiculo:
            filtros_texto.append(f'Vehículo: {tipo_vehiculo}')
        if tipo_documento:
            filtros_texto.append(f'Documento: {tipo_documento}')
        if start_date_str:
            filtros_texto.append(f'Desde: {start_date_str}')
        if end_date_str:
            filtros_texto.append(f'Hasta: {end_date_str}')

        fecha_generacion = datetime.now().strftime('%d/%m/%Y %H:%M')
        subtitulo = f'Generado: {fecha_generacion}'
        if filtros_texto:
            subtitulo += f'  |  Filtros: {", ".join(filtros_texto)}'

        ws.merge_cells('A2:T2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = subtitulo
        subtitle_cell.font = Font(name='Calibri', size=10, color='666666', italic=True)
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 22

        # Fila vacía de separación
        ws.row_dimensions[3].height = 8

        # Encabezados (fila 4)
        headers = [
            ('ID', 8),
            ('Placa', 12),
            ('Nombre Completo', 28),
            ('No. Documento', 16),
            ('Tipo Documento', 15),
            ('Teléfono', 14),
            ('Es Propietario', 14),
            ('Tipo Trámite', 15),
            ('Tipo Vehículo', 14),
            ('Marca', 14),
            ('Línea', 14),
            ('Modelo', 10),
            ('Clase', 14),
            ('Tipo Servicio', 14),
            ('Color', 12),
            ('Precio de Ley', 14),
            ('Comisión', 14),
            ('Cliente', 22),
            ('Registrado por', 20),
            ('Fecha Registro', 18),
        ]

        header_row = 4
        for col_idx, (header_name, width) in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        ws.row_dimensions[header_row].height = 30

        # Mapeo de tipos de documento para legibilidad
        tipo_doc_display = {
            'C': 'CC', 'E': 'CE', 'N': 'NIT', 'P': 'Pasaporte',
            'T': 'TI', 'R': 'RC', 'D': 'DIE', 'S': 'Salvoconducto',
            'L': 'Licencia', 'PE': 'PEP', 'PT': 'PPT',
        }

        # Estilos alternados para filas
        even_fill = PatternFill(start_color='F5F9FF', end_color='F5F9FF', fill_type='solid')

        # Datos
        for row_idx, registro in enumerate(registros, start=header_row + 1):
            usuario_name = ''
            if registro.usuario:
                usuario_name = f"{registro.usuario.first_name} {registro.usuario.last_name}".strip()

            cliente_nombre = registro.cliente.nombre if registro.cliente else ''

            fecha_creacion = registro.created_at.replace(tzinfo=None) if registro.created_at else None

            row_data = [
                registro.id,
                registro.placa or '',
                registro.nombre_completo or '',
                registro.numero_documento or '',
                tipo_doc_display.get(registro.tipo_documento, registro.tipo_documento),
                registro.telefono or '',
                'Sí' if registro.es_propietario else 'No',
                registro.tipo_tramite or '',
                registro.tipo_vehiculo or '',
                registro.marca or '',
                registro.linea or '',
                registro.modelo or '',
                registro.clase or '',
                registro.tipo_servicio or '',
                registro.color or '',
                float(registro.precio_lay) if registro.precio_lay is not None else None,
                float(registro.comision) if registro.comision is not None else None,
                cliente_nombre,
                usuario_name,
                fecha_creacion,
            ]

            is_even = (row_idx - header_row) % 2 == 0

            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = cell_alignment
                cell.font = Font(name='Calibri', size=10)
                if is_even:
                    cell.fill = even_fill
                # Formato de fecha
                if col_idx == len(row_data) and isinstance(value, datetime):
                    cell.number_format = date_format

        # Total de registros al final
        total_row = header_row + registros.count() + 2
        ws.cell(row=total_row, column=1, value='Total de registros:').font = Font(
            name='Calibri', bold=True, size=11
        )
        ws.cell(row=total_row, column=2, value=registros.count()).font = Font(
            name='Calibri', bold=True, size=11, color='1976D2'
        )

        # Congelar paneles (encabezados siempre visibles)
        ws.freeze_panes = f'A{header_row + 1}'

        # Autofiltro en encabezados
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f'A{header_row}:{last_col}{header_row + registros.count()}'

        # Generar archivo
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Nombre del archivo
        fecha_archivo = datetime.now().strftime('%Y%m%d_%H%M')
        filename = f'base_de_datos_registros_{fecha_archivo}.xlsx'

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        return Response(
            {"error": f"Error al exportar registros: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated, RolePermission(['admin', 'SuperAdmin']), ModulePermission('base_de_datos', 'view')])
def registro_history(request, pk):
    """Obtener el historial de cambios de un registro de vehículo"""
    try:
        registro = get_object_or_404(RegistroVehiculo, pk=pk)
        history = registro.history.all()

        # Paginación
        page_size_param = request.query_params.get('page_size', 10)
        try:
            page_size_int = int(page_size_param)
        except (ValueError, TypeError):
            page_size_int = 10

        paginator = PageNumberPagination()
        paginator.page_size = page_size_int
        paginated_history = paginator.paginate_queryset(history, request)

        data = []
        for h in paginated_history:
            data.append({
                'history_id': h.history_id,
                'history_date': h.history_date,
                'history_type': h.history_type,
                'history_type_display': h.get_history_type_display(),
                'history_user': {
                    'id': h.history_user.id,
                    'name': f"{h.history_user.first_name} {h.history_user.last_name}".strip()
                } if h.history_user else None,
                'tipo_tramite': h.tipo_tramite,
                'tipo_vehiculo': h.tipo_vehiculo,
                'es_propietario': h.es_propietario,
                'tipo_documento': h.tipo_documento,
                'numero_documento': h.numero_documento,
                'nombre_completo': h.nombre_completo,
                'telefono': h.telefono,
                'placa': h.placa,
                'clase': h.clase,
                'clasificacion': h.clasificacion,
                'tipo_servicio': h.tipo_servicio,
                'tipo_carroceria': h.tipo_carroceria,
                'vin': h.vin,
                'num_motor': h.num_motor,
                'num_chasis': h.num_chasis,
                'marca': h.marca,
                'linea': h.linea,
                'modelo': h.modelo,
                'cilindraje': h.cilindraje,
                'color': h.color,
                'organismo_transito': h.organismo_transito,
            })

        return paginator.get_paginated_response(data)

    except Exception as e:
        return Response(
            {"error": f"Error al obtener historial: {str(e)}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
